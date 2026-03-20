import os
import re
import csv
import tempfile
import asyncio
from datetime import datetime, date
from io import BytesIO
from typing import Dict, List, Tuple, Optional, Any

from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile
from openpyxl import load_workbook, Workbook
from openpyxl.utils.datetime import from_excel

# ========= ENV =========
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()

# contoh Railway:
# ADMIN_IDS=5397964203,123456789,987654321
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "").strip()

if not BOT_TOKEN:
    raise SystemExit("Missing BOT_TOKEN env")
if not ADMIN_IDS_RAW:
    raise SystemExit("Missing ADMIN_IDS env (comma separated digits)")

ADMIN_IDS: set[int] = set()
for x in re.split(r"[,\s]+", ADMIN_IDS_RAW):
    x = x.strip()
    if not x:
        continue
    if not x.isdigit():
        raise SystemExit(f"Invalid ADMIN_IDS value: {x} (must be digits)")
    ADMIN_IDS.add(int(x))

if not ADMIN_IDS:
    raise SystemExit("ADMIN_IDS is empty")

bot = Bot(BOT_TOKEN)
dp = Dispatcher()

# session per admin
# awaiting: None | "idsfile" | "file"
SESS: Dict[int, Dict[str, Any]] = {}

RE_SPLIT = re.compile(r"[,\s]+")
DEPOSIT_KEYWORDS = ("deposit", "depo", "topup", "top up")

# alias header biar lebih fleksibel
DATE_ALIASES = {"date", "tanggal", "tgl", "datetime", "created_at", "waktu"}
INFO_ALIASES = {"info", "type", "jenis", "keterangan", "desc", "description", "remark"}
TO_ALIASES = {"to", "userid", "user_id", "id", "username", "user", "member", "account"}
COIN_ALIASES = {"coin", "amount", "nominal", "total", "value", "koin"}


def is_admin(m: Message) -> bool:
    u = m.from_user
    return bool(u and u.id in ADMIN_IDS)


async def deny_if_not_admin(m: Message) -> bool:
    if not is_admin(m):
        await m.answer("⛔ Akses ditolak. Bot ini khusus admin.")
        return True
    return False


def norm_text(s: Any) -> str:
    return str(s or "").strip()


def norm_id(s: Any) -> str:
    return norm_text(s).lower()


def fmt(n: int) -> str:
    try:
        return f"{int(n):,}"
    except Exception:
        return str(n)


def normalize_header_name(s: Any) -> str:
    s = norm_text(s).lower()
    s = re.sub(r"[^a-z0-9_ ]+", "", s)
    s = re.sub(r"\s+", "_", s).strip("_")
    return s


def find_first_key(keys: set[str], aliases: set[str]) -> Optional[str]:
    for a in aliases:
        if a in keys:
            return a
    return None


def parse_date_any(s: str) -> Optional[datetime]:
    s = norm_text(s)
    if not s:
        return None

    fmts = [
        "%d/%m/%y %H.%M",
        "%d/%m/%y %H:%M",
        "%d/%m/%Y %H.%M",
        "%d/%m/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
    ]
    for f in fmts:
        try:
            return datetime.strptime(s, f)
        except Exception:
            pass

    s2 = s.replace(".", ":")
    if s2 != s:
        for f in [
            "%d/%m/%y %H:%M",
            "%d/%m/%Y %H:%M",
            "%m/%d/%Y %H:%M",
        ]:
            try:
                return datetime.strptime(s2, f)
            except Exception:
                pass
    return None


def parse_excel_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, (int, float)):
        try:
            dt = from_excel(value)
            if isinstance(dt, datetime):
                return dt
            if isinstance(dt, date):
                return datetime.combine(dt, datetime.min.time())
        except Exception:
            return None

    return parse_date_any(str(value))


def to_int_coin(x: Any) -> int:
    try:
        if x is None:
            return 0

        if isinstance(x, str):
            xs = x.strip()
            if not xs:
                return 0
            xs = xs.replace(",", "").replace(" ", "")
            xs = re.sub(r"[^\d.\-eE+]", "", xs)
            if not xs:
                return 0
            return int(float(xs))

        if isinstance(x, (int, float)):
            return int(float(x))
    except Exception:
        return 0

    return 0


# =======================
# Load ID file (TXT/CSV/XLSX)
# =======================

def load_ids_from_txt(path: str) -> set[str]:
    ids: set[str] = set()
    with open(path, "r", encoding="utf-8-sig", errors="ignore") as f:
        for line in f:
            parts = RE_SPLIT.split(line.strip())
            for p in parts:
                s = norm_id(p)
                if s and s not in {"to", "id", "userid", "user_id", "none", "null"}:
                    ids.add(s)
    return ids


def load_ids_from_csv(path: str) -> set[str]:
    ids: set[str] = set()

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except Exception:
            dialect = csv.excel

        reader = csv.reader(f, dialect=dialect)
        rows = list(reader)

    if not rows:
        return ids

    first = rows[0]
    header = [normalize_header_name(c) for c in first]

    preferred_cols = ["to", "userid", "user_id", "id", "username", "user", "member", "account"]
    col_idx = None
    for col in preferred_cols:
        if col in header:
            col_idx = header.index(col)
            break

    data_rows = rows[1:] if col_idx is not None else rows

    for r in data_rows:
        if not r:
            continue
        idx = col_idx if (col_idx is not None and col_idx < len(r)) else 0
        val = r[idx] if idx < len(r) else ""
        s = norm_id(val)
        if s and s not in {"to", "id", "userid", "user_id", "none", "null"}:
            ids.add(s)

    return ids


def load_ids_from_xlsx(path: str) -> set[str]:
    ids: set[str] = set()
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ids

    first_row = rows[0]
    headers = [normalize_header_name(c) for c in first_row]

    preferred_cols = ["to", "userid", "user_id", "id", "username", "user", "member", "account"]
    col_idx = None
    for col in preferred_cols:
        if col in headers:
            col_idx = headers.index(col)
            break

    data_rows = rows[1:] if col_idx is not None else rows

    for row in data_rows:
        if not row:
            continue
        idx = col_idx if (col_idx is not None and col_idx < len(row)) else 0
        val = row[idx] if idx < len(row) else None
        s = norm_id(val)
        if s and s not in {"to", "id", "userid", "user_id", "none", "null"}:
            ids.add(s)

    return ids


def load_id_file(path: str) -> set[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".txt":
        return load_ids_from_txt(path)
    if ext == ".csv":
        return load_ids_from_csv(path)
    if ext == ".xlsx":
        return load_ids_from_xlsx(path)
    raise ValueError("File ID harus .txt / .csv / .xlsx")


# =======================
# Load history (CSV/XLSX)
# =======================

def load_rows_from_csv(path: str) -> List[dict]:
    rows: List[dict] = []

    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(8192)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
        except Exception:
            dialect = csv.excel

        reader = csv.DictReader(f, dialect=dialect)

        if reader.fieldnames:
            reader.fieldnames = [normalize_header_name(fn) for fn in reader.fieldnames]

        for r in reader:
            rr = {}
            for k, v in r.items():
                if k is None:
                    continue
                kk = normalize_header_name(k)
                rr[kk] = v.strip() if isinstance(v, str) else v
            rows.append(rr)

    return rows


def load_rows_from_xlsx(path: str) -> List[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    headers = [normalize_header_name(c) for c in raw_rows[0]]
    rows: List[dict] = []

    for row in raw_rows[1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        rr = {}
        for h, cell in zip(headers, row):
            if not h:
                continue
            rr[h] = cell.strip() if isinstance(cell, str) else cell
        rows.append(rr)

    return rows


def load_history(path: str) -> List[dict]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return load_rows_from_csv(path)
    if ext == ".xlsx":
        return load_rows_from_xlsx(path)
    raise ValueError("History harus CSV atau XLSX")


def compute_report(
    rows: List[dict],
    ids: set[str],
    requested_day: Optional[date]
) -> Tuple[date, List[Tuple[str, int, int]]]:
    if not rows:
        raise ValueError("File history kosong atau tidak ada data yang terbaca.")

    keys = set(rows[0].keys())

    date_key = find_first_key(keys, DATE_ALIASES)
    info_key = find_first_key(keys, INFO_ALIASES)
    to_key = find_first_key(keys, TO_ALIASES)
    coin_key = find_first_key(keys, COIN_ALIASES)

    missing = []
    if not date_key:
        missing.append("date/tanggal")
    if not info_key:
        missing.append("info/keterangan")
    if not to_key:
        missing.append("to/user/id")
    if not coin_key:
        missing.append("coin/amount/nominal")

    if missing:
        raise ValueError(
            f"Kolom wajib tidak ditemukan: {', '.join(missing)}\n"
            f"Kolom yang terbaca di file: {', '.join(sorted(keys))}"
        )

    parsed: List[Tuple[date, str, int]] = []

    for r in rows:
        info = norm_text(r.get(info_key, "")).lower()
        if not any(k in info for k in DEPOSIT_KEYWORDS):
            continue

        dtp = parse_excel_datetime(r.get(date_key))
        if not dtp:
            continue

        to_id = norm_id(r.get(to_key, ""))
        if not to_id or to_id not in ids:
            continue

        coin = to_int_coin(r.get(coin_key, 0))
        parsed.append((dtp.date(), to_id, coin))

    if not parsed:
        raise ValueError("Tidak ada data deposit yang cocok dengan ID new member di file history.")

    available_days = sorted({d for d, _, _ in parsed})
    target_day = requested_day if (requested_day in available_days) else available_days[-1]

    agg: Dict[str, Dict[str, int]] = {}
    for d, to_id, coin in parsed:
        if d != target_day:
            continue
        if to_id not in agg:
            agg[to_id] = {"count": 0, "sum": 0}
        agg[to_id]["count"] += 1
        agg[to_id]["sum"] += coin

    if not agg:
        raise ValueError(f"Tidak ada deposit untuk tanggal {target_day} pada ID yang cocok.")

    out = [(k, v["count"], v["sum"]) for k, v in agg.items()]
    out.sort(key=lambda x: (x[2], x[1], x[0]), reverse=True)
    return target_day, out


def make_excel_bytes(target_day: date, items: List[Tuple[str, int, int]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(["TO", "Deposit Count", "Total Coin", "Tanggal"])

    for to_id, cnt, total in items:
        ws.append([to_id, cnt, total, str(target_day)])

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ================== COMMANDS ==================

@dp.message(Command("start"))
async def start(m: Message):
    if await deny_if_not_admin(m):
        return

    SESS[m.from_user.id] = {"ids": set(), "target_date": None, "awaiting": None}

    await m.answer(
        "✅ Bot Cek New Member siap ketua\n\n"
        "Perintah:\n"
        "/ceknew = mulai (upload file ID dulu)\n"
        "/tanggal YYYY-MM-DD = set tanggal (opsional)\n"
        "/reset = hapus session\n"
        "/adminlist = lihat daftar admin\n\n"
        "Alur:\n"
        "1) /ceknew\n"
        "2) upload file ID (.txt/.csv/.xlsx)\n"
        "3) upload history koin (.csv/.xlsx)\n"
    )


@dp.message(Command("reset"))
async def reset(m: Message):
    if await deny_if_not_admin(m):
        return

    SESS[m.from_user.id] = {"ids": set(), "target_date": None, "awaiting": None}
    await m.answer("♻️ Session direset. Ketik /ceknew untuk mulai lagi.")


@dp.message(Command("tanggal"))
async def tanggal(m: Message):
    if await deny_if_not_admin(m):
        return

    parts = norm_text(m.text).split(maxsplit=1)
    if len(parts) < 2:
        await m.answer("Format: /tanggal 2026-01-13")
        return

    try:
        d = datetime.strptime(parts[1].strip(), "%Y-%m-%d").date()
    except Exception:
        await m.answer("Format tanggal salah. Contoh: /tanggal 2026-01-13")
        return

    sess = SESS.setdefault(m.from_user.id, {"ids": set(), "target_date": None, "awaiting": None})
    sess["target_date"] = d
    await m.answer(f"📅 OK. Target tanggal diset: {d} (kalau tidak ada di file, bot pakai tanggal terbaru).")


@dp.message(Command("ceknew"))
async def ceknew(m: Message):
    if await deny_if_not_admin(m):
        return

    sess = SESS.setdefault(m.from_user.id, {"ids": set(), "target_date": None, "awaiting": None})
    sess["ids"] = set()
    sess["awaiting"] = "idsfile"

    await m.answer(
        "Upload file ID new member dulu ✅\n"
        "Format: .txt / .csv / .xlsx\n"
        "- TXT: boleh 1 baris 1 ID atau dipisah spasi/koma\n"
        "- CSV/XLSX: kolom TO / ID / USERID (atau kolom A)\n\n"
        "Setelah itu upload history koin."
    )


@dp.message(Command("adminlist"))
async def adminlist(m: Message):
    if await deny_if_not_admin(m):
        return

    await m.answer("✅ Admin IDs:\n" + "\n".join(str(i) for i in sorted(ADMIN_IDS)))


# ================== DOCUMENT HANDLER ==================

@dp.message(F.document)
async def handle_doc(m: Message):
    if await deny_if_not_admin(m):
        return

    sess = SESS.setdefault(m.from_user.id, {"ids": set(), "target_date": None, "awaiting": None})
    doc = m.document

    if not doc:
        await m.answer("❌ Dokumen tidak ditemukan.")
        return

    fname = norm_text(doc.file_name) or "file"
    ext = os.path.splitext(fname)[1].lower()

    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, fname)

        try:
            await bot.download(doc, destination=path)
        except Exception as e:
            await m.answer(f"❌ Gagal download file: {e}")
            return

        # 1) upload ID file
        if sess.get("awaiting") == "idsfile":
            if ext not in [".txt", ".csv", ".xlsx"]:
                await m.answer("❌ File ID harus .txt / .csv / .xlsx")
                return

            try:
                ids = load_id_file(path)
            except Exception as e:
                await m.answer(f"❌ Gagal baca file ID: {e}")
                return

            if not ids:
                await m.answer("❌ ID kosong. Pastikan file berisi ID.")
                return

            sess["ids"] = ids
            sess["awaiting"] = "file"

            await m.answer(
                f"✅ ID terkumpul: {fmt(len(ids))}\n"
                "Sekarang upload history koin (CSV/XLSX)."
            )
            return

        # 2) upload history file
        if sess.get("awaiting") == "file":
            if ext not in [".csv", ".xlsx"]:
                await m.answer("❌ History harus CSV atau XLSX")
                return

            if not sess.get("ids"):
                await m.answer("❌ ID belum ada. Ketik /ceknew lalu upload file ID dulu.")
                return

            try:
                rows = load_history(path)
                target_day, items = compute_report(rows, sess["ids"], sess.get("target_date"))
            except Exception as e:
                await m.answer(f"❌ Gagal proses history: {e}")
                return

            total_member_dp = len(items)
            total_dp_count = sum(cnt for _, cnt, _ in items)
            total_coin = sum(total for _, _, total in items)

            lines = []
            for i, (to_id, cnt, total) in enumerate(items[:15], start=1):
                lines.append(f"{i}. {to_id} | {fmt(cnt)}x | {fmt(total)}")

            msg = (
                f"📊 Report New Member (Tanggal {target_day})\n\n"
                f"📌 Total ID input: {fmt(len(sess['ids']))}\n"
                f"✅ Member yang deposit: {fmt(total_member_dp)}\n"
                f"🧾 Total deposit count (all new member): {fmt(total_dp_count)}\n"
                f"🪙 Total coin deposit (all new member): {fmt(total_coin)}\n\n"
                f"🏆 Top 15:\n"
                + ("\n".join(lines) if lines else "-")
            )

            await m.answer(msg)

            try:
                xbytes = make_excel_bytes(target_day, items)
                outname = f"report_new_member_{target_day}.xlsx"
                await m.answer_document(BufferedInputFile(xbytes, filename=outname))
            except Exception as e:
                await m.answer(f"⚠️ Report tampil, tapi gagal kirim Excel: {e}")

            # reset session setelah selesai
            sess["ids"] = set()
            sess["target_date"] = None
            sess["awaiting"] = None
            return

    await m.answer("Ketik /ceknew dulu, lalu upload file ID (.txt/.csv/.xlsx).")


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
